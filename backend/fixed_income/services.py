"""
Services for fixed income portfolio import and management.
"""
import re
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Tuple
import openpyxl
from django.utils import timezone
from .models import FixedIncomePosition, TesouroDiretoPosition, InvestmentFund
from configuration.models import InvestmentType, InvestmentSubType


class PortfolioExcelImportService:
    """Service for importing portfolio positions from Excel files."""
    
    @staticmethod
    def parse_currency(value: str) -> Decimal:
        """Parse Brazilian currency format (R$ X.XXX,XX) to Decimal."""
        if not value or value == '-':
            return Decimal('0.00')
        
        # Remove R$ and spaces
        value = str(value).replace('R$', '').strip()
        
        # Handle comma as decimal separator
        if ',' in value:
            # Replace dots (thousands) and comma (decimal)
            value = value.replace('.', '').replace(',', '.')
        else:
            # No comma, assume it's already in standard format
            pass
        
        try:
            return Decimal(value)
        except (InvalidOperation, ValueError):
            return Decimal('0.00')
    
    @staticmethod
    def parse_date(date_str: str) -> Optional[datetime]:
        """Parse date string in DD/MM/YYYY format."""
        if not date_str or date_str == '-':
            return None
        
        try:
            # Handle Excel date format or string format
            if isinstance(date_str, datetime):
                return date_str
            elif isinstance(date_str, str):
                # Try DD/MM/YYYY format
                parts = date_str.split('/')
                if len(parts) == 3:
                    day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                    return datetime(year, month, day)
        except (ValueError, AttributeError):
            pass
        
        return None
    
    @staticmethod
    def parse_percentage(value: str) -> Optional[str]:
        """Parse percentage value (e.g., '+9,00%')."""
        if not value or value == '-':
            return None
        return str(value).strip()
    
    @staticmethod
    def parse_quantity(value: str) -> Decimal:
        """Parse quantity value."""
        if not value or value == '-':
            return Decimal('0.00')
        
        # Handle comma as decimal separator
        value = str(value).replace(',', '.')
        try:
            return Decimal(value)
        except (InvalidOperation, ValueError):
            return Decimal('0.00')
    
    @staticmethod
    def detect_section(row: List) -> Optional[str]:
        """Detect which section a row belongs to (Ações, Renda Fixa, Tesouro Direto, Fundos de Investimento)."""
        if not row or len(row) == 0:
            return None
        
        first_cell = str(row[0]).upper() if row[0] else ''
        
        # Skip section headers that are proventos/dividendos/distribuições (these are not actual sections)
        excluded_keywords = [
            'PROVENTOS', 'DIVIDENDOS', 'DISTRIBUIÇÕES', 'DISTRIBUICOES',
            'RENDIMENTOS', 'JUROS', 'AMORTIZAÇÃO', 'AMORTIZACAO'
        ]
        if any(keyword in first_cell for keyword in excluded_keywords):
            return None
        
        # Check for Fundos de Investimento first (before Renda Fixa, to avoid false matches)
        if 'FUNDOS DE INVESTIMENTO' in first_cell or 'FUNDO DE INVESTIMENTO' in first_cell:
            return 'fundos_investimento'
        elif 'AÇÕES' in first_cell or '%|AÇÕES' in first_cell:
            return 'acoes'
        elif 'RENDA FIXA' in first_cell or '%|RENDA FIXA' in first_cell:
            # Avoid matching "Fundos de Renda Fixa" subsections
            if 'FUNDOS' not in first_cell and 'FUNDO' not in first_cell:
                return 'renda_fixa'
        elif 'TESOURO' in first_cell or 'TESOURO DIRETO' in first_cell or '%|TESOURO' in first_cell or 'TESOURO DIRETO' in first_cell:
            return 'tesouro_direto'
        
        return None
    
    @staticmethod
    def detect_tesouro_subsection(row: List) -> Optional[Tuple[str, Decimal]]:
        """Detect Tesouro Direto sub-section header (e.g., '15,2% | Pós-Fixado').
        Returns tuple of (sub_type_name, allocation_percentage) or None."""
        if not row or len(row) == 0:
            return None
        
        first_cell = str(row[0]).strip() if row[0] else ''
        
        # Pattern: "{percentage}% | {sub-type}"
        # Examples: "15,2% | Pós-Fixado", "14,3% | Inflação", "4,6% | Prefixado"
        # Also handles: "15,20%|Pós-Fixada", "4,50%|Prefixada"
        pattern = r'([\d,]+)\s*%\s*\|\s*(.+)'
        match = re.search(pattern, first_cell, re.IGNORECASE)
        
        if match:
            percentage_str = match.group(1).replace(',', '.')
            sub_type_name = match.group(2).strip()
            
            # Normalize sub-type names (handle variations)
            sub_type_name_upper = sub_type_name.upper()
            if 'PÓS-FIXAD' in sub_type_name_upper or 'POS-FIXAD' in sub_type_name_upper:
                sub_type_name = 'Pós-Fixado'
            elif 'PREFIXAD' in sub_type_name_upper:
                sub_type_name = 'Prefixado'
            elif 'INFLA' in sub_type_name_upper:
                sub_type_name = 'Inflação'
            
            try:
                percentage = Decimal(percentage_str)
                return (sub_type_name, percentage)
            except (InvalidOperation, ValueError):
                pass
        
        return None
    
    @staticmethod
    def detect_fund_type_subsection(row: List) -> Optional[str]:
        """Detect Investment Fund sub-section header (e.g., '0,54%|Fundos de Renda Fixa Pós-Fixado').
        Returns fund type string or None."""
        if not row or len(row) == 0:
            return None
        
        first_cell = str(row[0]).strip() if row[0] else ''
        
        # Pattern: "{percentage}%|{fund-type}"
        # Examples: "0,54%|Fundos de Renda Fixa Pós-Fixado"
        pattern = r'[\d,]+\s*%\s*\|\s*(.+)'
        match = re.search(pattern, first_cell, re.IGNORECASE)
        
        if match:
            fund_type_str = match.group(1).strip()
            fund_type_upper = fund_type_str.upper()
            
            # Determine fund type based on description
            if 'PÓS-FIXADO' in fund_type_upper or 'POS-FIXADO' in fund_type_upper or 'PÓS FIXADO' in fund_type_upper:
                return 'RF_POS'
            elif 'PREFIXADO' in fund_type_upper or 'PRE-FIXADO' in fund_type_upper or 'PRÉ-FIXADO' in fund_type_upper:
                return 'RF_PRE'
            elif 'IPCA' in fund_type_upper or 'INFLAÇÃO' in fund_type_upper:
                return 'RF_IPCA'
            elif 'MULTIMERCADO' in fund_type_upper:
                return 'MULTI'
            elif 'AÇÕES' in fund_type_upper or 'ACOES' in fund_type_upper:
                return 'ACOES'
            elif 'CÂMBIO' in fund_type_upper or 'CAMBIO' in fund_type_upper:
                return 'CAMBIO'
            else:
                return 'RF_POS'  # Default to Pós-Fixado for Renda Fixa funds
        
        return None
    
    @staticmethod
    def is_header_row(row: List) -> bool:
        """Check if a row is a header row."""
        if not row:
            return False
        
        header_keywords = ['ATIVO', 'QTD', 'PREÇO', 'POSIÇÃO', 'APLICAÇÃO', 'VENCIMENTO', 'ALOCAÇÃO', 'DISPONÍVEL', 'TOTAL APLICADO']
        first_cell = str(row[0]).upper() if row[0] else ''
        
        return any(keyword in first_cell for keyword in header_keywords)
    
    @staticmethod
    def get_tesouro_subtype_from_bond(titulo_name: str) -> Optional[str]:
        """Determine Tesouro Direto sub-type from bond title.
        Returns sub-type name: 'Pós-Fixado', 'Inflação', or 'Prefixado'."""
        titulo_upper = titulo_name.upper().strip()
        
        # LFT (Letra Financeira do Tesouro) → Pós-Fixado
        if 'LFT' in titulo_upper:
            return 'Pós-Fixado'
        # NTNB (Nota do Tesouro Nacional - B) → Inflação
        # Also handles "NTNB PRINC" format
        elif 'NTNB' in titulo_upper:
            return 'Inflação'
        # LTN (Letra do Tesouro Nacional) → Prefixado
        elif 'LTN' in titulo_upper:
            return 'Prefixado'
        # NTN-B (alternative notation) → Inflação
        elif 'NTN-B' in titulo_upper or 'NTN B' in titulo_upper:
            return 'Inflação'
        
        return None
    
    @staticmethod
    def extract_cdb_from_row(row: List, user_id: str, current_section: str) -> Optional[Dict]:
        """Extract CDB data from a row."""
        if not row or len(row) < 3:
            return None
        
        asset_name = str(row[0]).strip() if row[0] else ''
        
        # Skip entries that are not actual investments (proventos, dividendos, distribuições, etc.)
        asset_name_upper = asset_name.upper()
        excluded_keywords = [
            'PROVENTOS', 'DIVIDENDOS', 'DISTRIBUIÇÕES', 'DISTRIBUICOES',
            'RENDIMENTOS', 'JUROS', 'AMORTIZAÇÃO', 'AMORTIZACAO',
            'RESGATE', 'LIQUIDAÇÃO', 'LIQUIDACAO'
        ]
        
        if any(keyword in asset_name_upper for keyword in excluded_keywords):
            return None
        
        # Check if it's a CDB
        if 'CDB' not in asset_name.upper():
            return None
        
        # CDB rows in Excel format:
        # [Asset Name, Application Date, Grace Period End, Maturity Date, Rate, Quantity, Guarantee, Applied Value, Position Value, Net Value, ...]
        
        try:
            application_date = PortfolioExcelImportService.parse_date(row[1])
            grace_period_end = PortfolioExcelImportService.parse_date(row[2]) if len(row) > 2 else None
            maturity_date = PortfolioExcelImportService.parse_date(row[3]) if len(row) > 3 else None
            
            # If maturity_date is None, try to extract from asset_name (e.g., "CDB BANCO MASTER S/A - DEZ/2026")
            if not maturity_date:
                maturity_date = PortfolioExcelImportService._extract_date_from_name(asset_name)
            
            rate = PortfolioExcelImportService.parse_percentage(row[4]) if len(row) > 4 else None
            quantity = PortfolioExcelImportService.parse_quantity(row[5]) if len(row) > 5 else Decimal('0.00')
            guarantee_quantity = PortfolioExcelImportService.parse_quantity(row[6]) if len(row) > 6 else Decimal('0.00')
            applied_value = PortfolioExcelImportService.parse_currency(row[7]) if len(row) > 7 else Decimal('0.00')
            position_value = PortfolioExcelImportService.parse_currency(row[8]) if len(row) > 8 else Decimal('0.00')
            net_value = PortfolioExcelImportService.parse_currency(row[9]) if len(row) > 9 else Decimal('0.00')
            
            # Extract asset code from asset_name or generate one
            asset_code = PortfolioExcelImportService._extract_asset_code(asset_name)
            
            # Calculate yields
            gross_yield = position_value - applied_value if position_value and applied_value else Decimal('0.00')
            net_yield = net_value - applied_value if net_value and applied_value else Decimal('0.00')
            income_tax = gross_yield - net_yield if gross_yield and net_yield else Decimal('0.00')
            
            if not application_date:
                return None
            
            return {
                'user_id': user_id,
                'asset_name': asset_name,
                'asset_code': asset_code,
                'application_date': application_date.date() if isinstance(application_date, datetime) else application_date,
                'grace_period_end': grace_period_end.date() if grace_period_end and isinstance(grace_period_end, datetime) else grace_period_end,
                'maturity_date': maturity_date.date() if maturity_date and isinstance(maturity_date, datetime) else maturity_date,
                'rate': rate,
                'quantity': quantity,
                'available_quantity': quantity,  # Assume same as quantity initially
                'guarantee_quantity': guarantee_quantity,
                'applied_value': applied_value,
                'position_value': position_value,
                'net_value': net_value,
                'gross_yield': gross_yield,
                'net_yield': net_yield,
                'income_tax': income_tax,
                'iof': Decimal('0.00'),  # Default, can be updated later
                'source': 'Excel Import',
                'import_date': timezone.now(),
            }
        except Exception as e:
            print(f"Error extracting CDB from row: {e}")
            return None
    
    @staticmethod
    def _extract_date_from_name(asset_name: str) -> Optional[datetime]:
        """Extract date from asset name (e.g., 'CDB BANCO MASTER S/A - DEZ/2026')."""
        # Look for month/year pattern (e.g., DEZ/2026, SET/2031)
        month_map = {
            'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4, 'MAI': 5, 'JUN': 6,
            'JUL': 7, 'AGO': 8, 'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12
        }
        
        pattern = r'([A-Z]{3})/(\d{4})'
        match = re.search(pattern, asset_name.upper())
        
        if match:
            month_str, year_str = match.groups()
            if month_str in month_map:
                try:
                    return datetime(int(year_str), month_map[month_str], 1)
                except ValueError:
                    pass
        
        return None
    
    @staticmethod
    def _extract_asset_code(asset_name: str) -> str:
        """Extract or generate asset code from asset name."""
        # Try to find a code pattern in the name
        # For now, use a simplified version of the name as code
        # In real scenarios, this might come from a separate column
        code = asset_name.replace(' ', '_').upper()[:50]
        return code
    
    @staticmethod
    def extract_tesouro_from_row(row: List, user_id: str, sub_type_name: Optional[str] = None) -> Optional[Dict]:
        """Extract Tesouro Direto data from a row.
        Actual Excel format based on portfolio export:
        [Título, Posição, % Alocação, Total aplicado, Qtd., Disponível, Vencimento]
        Column mapping:
        - Row[0]: Título name (e.g., "LFT mar/2029", "LTN jan/2027", "NTNB PRINC mai/2029")
        - Row[1]: Posição (position_value) - "R$ 29.013,40"
        - Row[2]: % Alocação (allocation percentage)
        - Row[3]: Total aplicado (applied_value) - "R$ 25.125,32"
        - Row[4]: Qtd. (quantity) - "1,63"
        - Row[5]: Disponível (available_quantity) - "1,63"
        - Row[6]: Vencimento (maturity_date) - "01/03/2029"
        """
        if not row or len(row) < 2:
            return None
        
        titulo_name = str(row[0]).strip() if row[0] else ''
        
        # Check if it's a Tesouro Direto bond (LFT, NTNB, LTN, etc.)
        if not any(bond_type in titulo_name.upper() for bond_type in ['LFT', 'NTNB', 'LTN', 'NTN-B', 'NTN B', 'TESOURO']):
            return None
        
        try:
            # Parse columns according to actual Excel format:
            # [Título, Posição, % Alocação, Total aplicado, Qtd., Disponível, Vencimento]
            
            # Row[1]: Posição (position_value) - current position value
            position_value = PortfolioExcelImportService.parse_currency(row[1]) if len(row) > 1 and row[1] is not None else Decimal('0.00')
            
            # Row[3]: Total aplicado (applied_value) - original investment amount
            applied_value = PortfolioExcelImportService.parse_currency(row[3]) if len(row) > 3 and row[3] is not None else Decimal('0.00')
            
            # Row[4]: Qtd. (quantity) - number of bonds
            if len(row) > 4 and row[4] is not None:
                quantity = PortfolioExcelImportService.parse_quantity(str(row[4]))
            else:
                quantity = Decimal('0.00')
            
            # Row[5]: Disponível (available_quantity)
            if len(row) > 5 and row[5] is not None:
                available_quantity = PortfolioExcelImportService.parse_quantity(str(row[5]))
            else:
                available_quantity = quantity  # Default to quantity if not specified
            
            # Row[6]: Vencimento (maturity_date)
            vencimento = PortfolioExcelImportService.parse_date(row[6]) if len(row) > 6 and row[6] is not None else None
            
            # Price is not directly in the Excel for this format
            # We can calculate it from applied_value / quantity if both are available
            if quantity > 0 and applied_value > 0:
                price = applied_value / quantity
            else:
                price = Decimal('0.00')
            
            # Garantia (guarantee_quantity) is not in this format, default to 0
            guarantee_quantity = Decimal('0.00')
            
            if not vencimento:
                return None
            
            # Convert vencimento to date object
            from datetime import date as date_class
            vencimento_date = vencimento.date() if isinstance(vencimento, datetime) else vencimento
            
            # Determine sub-type if not provided
            if not sub_type_name:
                sub_type_name = PortfolioExcelImportService.get_tesouro_subtype_from_bond(titulo_name)
            
            # Build full asset name - título name from Excel already includes maturity date
            # Format from Excel: "LFT mar/2029", "LTN jan/2027", "NTNB PRINC mai/2029"
            # So we can use it directly
            full_titulo_name = titulo_name.strip()
            
            # Generate asset code from título name and maturity date
            # Extract just the bond type (LFT, LTN, NTNB PRINC, etc.) without the month/year suffix
            # The título format is like "LFT mar/2029" or "NTNB PRINC mai/2029"
            # Split and take only the parts before the date (month/year format)
            parts = titulo_name.strip().split()
            bond_type_parts = []
            for part in parts:
                # Stop at parts that look like dates (month/year format like "mar/2029" or "mai/2029")
                if '/' in part and len(part.split('/')) == 2:
                    # Check if it's a month/year format (has lowercase letters + numbers)
                    month_part, year_part = part.split('/')
                    if month_part.isalpha() and year_part.isdigit():
                        break
                bond_type_parts.append(part)
            
            # Join bond type parts (e.g., "NTNB PRINC" or "LFT")
            bond_type = ' '.join(bond_type_parts).strip() if bond_type_parts else titulo_name.split()[0] if titulo_name.split() else 'TESOURO'
            
            # Generate asset code using bond type and maturity date
            if isinstance(vencimento_date, (datetime, date_class)):
                if isinstance(vencimento_date, datetime):
                    date_str = vencimento_date.date().strftime('%Y%m%d')
                else:
                    date_str = vencimento_date.strftime('%Y%m%d')
                asset_code = f"TESOURO_{bond_type.replace(' ', '_').upper()}_{date_str}"
            else:
                asset_code = f"TESOURO_{bond_type.replace(' ', '_').upper()}_{str(vencimento_date).replace('-', '')}"
            
            # Calculate yields
            gross_yield = position_value - applied_value if position_value and applied_value else Decimal('0.00')
            net_yield = gross_yield  # For Tesouro Direto, net is typically same as gross (taxes applied at sale)
            net_value = position_value  # Net value is same as position value for Tesouro Direto
            
            # For Tesouro positions, use a fixed application_date based on maturity year
            # to ensure consistency and prevent duplicates when re-importing
            # The bond itself is unique by título + maturity, not by when it was imported
            if isinstance(vencimento_date, datetime):
                maturity_year = vencimento_date.year
            elif isinstance(vencimento_date, date_class):
                maturity_year = vencimento_date.year
            else:
                maturity_year = timezone.now().year
            
            # Use January 1st of the maturity year as a fixed application_date
            # This ensures the same bond always has the same application_date
            application_date = datetime(maturity_year, 1, 1).date()
            
            return {
                'user_id': user_id,
                'asset_name': full_titulo_name,
                'asset_code': asset_code,
                'application_date': application_date,
                'maturity_date': vencimento_date,
                'price': price,
                'quantity': quantity,
                'available_quantity': available_quantity,
                'guarantee_quantity': guarantee_quantity,
                'applied_value': applied_value,
                'position_value': position_value,
                'net_value': net_value,
                'gross_yield': gross_yield,
                'net_yield': net_yield,
                'income_tax': Decimal('0.00'),  # Taxes applied at sale
                'iof': Decimal('0.00'),
                'source': 'Excel Import',
                'import_date': timezone.now(),
                'titulo_name': full_titulo_name,
                'vencimento': vencimento_date,
                'sub_type_name': sub_type_name,  # Store for later use in import
            }
        except Exception as e:
            print(f"Error extracting Tesouro from row: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    @staticmethod
    def extract_investment_fund_from_row(row: List, user_id: str, fund_type: str = 'RF_POS') -> Optional[Dict]:
        """Extract investment fund data from a row.
        Supports two formats:
        1. Old format: Fund Name | Data cota | Valor cota | Qtd. cotas | Em cotização | Posição | Valor líquido
        2. New format (Posição Detalhada): Fund Name | Posição | % Alocação | Rent. Líquida | Rent. Bruta | Valor aplicado | Valor líquido
        """
        try:
            if not row or len(row) < 7:
                return None
            
            fund_name = str(row[0]).strip() if row[0] else None
            if not fund_name or fund_name == '-':
                return None
            
            # Skip if it looks like a header
            header_keywords = ['DATA COTA', 'VALOR COTA', 'POSIÇÃO', 'FUNDO', 'TIPO', 'QTD', 'COTAS', 'ALOCAÇÃO', 'RENTABILIDADE', 'VALOR APLICADO', 'VALOR LÍQUIDO']
            if any(keyword in fund_name.upper() for keyword in header_keywords):
                return None
            
            # Skip entries that are not actual funds (proventos, dividendos, distribuições, etc.)
            fund_name_upper = fund_name.upper()
            excluded_keywords = [
                'PROVENTOS',
                'DIVIDENDOS',
                'DISTRIBUIÇÕES',
                'DISTRIBUICOES',
                'RENDIMENTOS',
                'JUROS',
                'AMORTIZAÇÃO',
                'AMORTIZACAO',
                'RESGATE',
                'LIQUIDAÇÃO',
                'LIQUIDACAO'
            ]
            
            for keyword in excluded_keywords:
                if keyword in fund_name_upper:
                    # Log that we're skipping this entry
                    print(f"Skipping non-fund entry: '{fund_name}' (contains '{keyword}')")
                    return None
            
            # Detect format by checking if row[1] looks like a date or a currency value
            # If row[1] is a date, use old format; if it's currency (Posição), use new format
            row1_str = str(row[1]).strip() if row[1] else ''
            is_new_format = False
            
            # Check if row[1] looks like currency (starts with R$ or is a number)
            if row1_str and ('R$' in row1_str or (row1_str.replace('.', '').replace(',', '').replace('-', '').strip().isdigit())):
                is_new_format = True
            
            if is_new_format:
                # New format: Fund Name | Posição | % Alocação | Rent. Líquida | Rent. Bruta | Valor aplicado | Valor líquido
                quota_date = None  # Not available in new format
                quota_value = None  # Not available in new format
                quota_quantity = Decimal('0.00')  # Not available in new format
                in_quotation = Decimal('0.00')  # Not available in new format
                
                # Parse position value (column 1)
                position_value = PortfolioExcelImportService.parse_currency(str(row[1])) if row[1] else Decimal('0.00')
                
                # Parse applied value (column 5)
                applied_value = PortfolioExcelImportService.parse_currency(str(row[5])) if len(row) > 5 and row[5] else Decimal('0.00')
                
                # Parse net value (column 6)
                net_value = PortfolioExcelImportService.parse_currency(str(row[6])) if len(row) > 6 and row[6] else Decimal('0.00')
            else:
                # Old format: Fund Name | Data cota | Valor cota | Qtd. cotas | Em cotização | Posição | Valor líquido
                # Parse quota date (column 1)
                quota_date = PortfolioExcelImportService.parse_date(str(row[1])) if row[1] else None
                
                # Parse quota value (column 2)
                quota_value = PortfolioExcelImportService.parse_currency(str(row[2])) if row[2] else Decimal('0.00')
                
                # Parse quota quantity (column 3)
                quota_quantity = PortfolioExcelImportService.parse_quantity(str(row[3])) if row[3] else Decimal('0.00')
                
                # Parse in quotation (column 4)
                in_quotation = PortfolioExcelImportService.parse_currency(str(row[4])) if row[4] else Decimal('0.00')
                
                # Parse position value (column 5)
                position_value = PortfolioExcelImportService.parse_currency(str(row[5])) if row[5] else Decimal('0.00')
                
                # Parse net value (column 6)
                net_value = PortfolioExcelImportService.parse_currency(str(row[6])) if row[6] else Decimal('0.00')
                
                # Calculate applied value (position - gain) for old format
                applied_value = position_value  # Simplified - could be more accurate with historical data
            
            # Calculate returns (if we have applied value)
            gross_return_percent = None
            net_return_percent = None
            if applied_value > 0 and net_value > 0:
                net_return_percent = ((net_value - applied_value) / applied_value) * 100
            
            return {
                'user_id': user_id,
                'fund_name': fund_name,
                'fund_type': fund_type,
                'quota_date': quota_date,
                'quota_value': quota_value,
                'quota_quantity': quota_quantity,
                'in_quotation': in_quotation,
                'position_value': position_value,
                'net_value': net_value,
                'applied_value': applied_value,
                'gross_return_percent': gross_return_percent,
                'net_return_percent': net_return_percent,
                'source': 'Excel Import',
                'import_date': timezone.now(),
            }
        except Exception as e:
            print(f"Error extracting Investment Fund from row: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    @staticmethod
    def extract_cash_balance(worksheet) -> Optional[Decimal]:
        """Extract Saldo Disponível (cash balance) from Excel.
        Looks for the value in row 4, column 3 (index 2) or row 61, column 1 (index 0)."""
        try:
            # Method 1: Check row 4, column 3 (after "Saldo Disponível" header in row 3)
            if worksheet.max_row >= 4:
                row_3 = list(worksheet.iter_rows())[2]  # Row 3 (0-indexed)
                row_4 = list(worksheet.iter_rows())[3]  # Row 4 (0-indexed)
                
                # Check if row 3 has "Saldo Disponível" header
                if row_3[2] and 'SALDO DISPON' in str(row_3[2].value).upper():
                    cash_value = row_4[2].value if len(row_4) > 2 else None
                    if cash_value is not None:
                        if isinstance(cash_value, (int, float)):
                            return Decimal(str(cash_value))
                        else:
                            return PortfolioExcelImportService.parse_currency(str(cash_value))
            
            # Method 2: Check row 61, column 1 (after "Saldo Disponível" header in row 60)
            if worksheet.max_row >= 61:
                row_60 = list(worksheet.iter_rows())[59]  # Row 60 (0-indexed)
                row_61 = list(worksheet.iter_rows())[60]  # Row 61 (0-indexed)
                
                # Check if row 60 has "Saldo Disponível" header
                if row_60[0] and 'SALDO DISPON' in str(row_60[0].value).upper():
                    cash_value = row_61[0].value if len(row_61) > 0 else None
                    if cash_value is not None:
                        if isinstance(cash_value, (int, float)):
                            return Decimal(str(cash_value))
                        else:
                            return PortfolioExcelImportService.parse_currency(str(cash_value))
        except Exception as e:
            print(f"Error extracting cash balance: {e}")
        
        return None
    
    @staticmethod
    def import_from_excel(file_path: str, user_id: str) -> Dict:
        """Import portfolio positions from Excel file."""
        results = {
            'created': 0,
            'updated': 0,
            'errors': [],
            'cdb_count': 0,
            'tesouro_count': 0,
            'caixa_count': 0,
            'debug_info': [],
        }
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            results['debug_info'].append(f"Worksheet: {worksheet.title}, Rows: {worksheet.max_row}, Cols: {worksheet.max_column}")
            
            current_section = None
            current_fund_type = None  # Track current fund type for investment funds
            
            # Get or create Renda Fixa investment type
            renda_fixa_type, _ = InvestmentType.objects.get_or_create(
                code='RENDA_FIXA',
                defaults={'name': 'Renda Fixa', 'display_order': 2}
            )
            
            # Get or create TESOURO_DIRETO as a sub-type of RENDA_FIXA (not as an investment type)
            tesouro_subtype, _ = InvestmentSubType.objects.get_or_create(
                investment_type=renda_fixa_type,
                code='TESOURO_DIRETO',
                defaults={'name': 'Tesouro Direto', 'display_order': 2, 'is_active': True}
            )
            
            # Get or create CDB_PREFIXADO sub-type for CDB positions
            cdb_prefixado_subtype, _ = InvestmentSubType.objects.get_or_create(
                investment_type=renda_fixa_type,
                code='CDB_PREFIXADO',
                defaults={'name': 'CDB Pré-fixado', 'display_order': 3, 'is_active': True}
            )
            
            # Get or create Investment Fund sub-types
            fund_posfixado_subtype, _ = InvestmentSubType.objects.get_or_create(
                investment_type=renda_fixa_type,
                code='FUNDO_RF_POS',
                defaults={'name': 'Fundos de Renda Fixa Pós-Fixado', 'display_order': 4, 'is_active': True}
            )
            fund_prefixado_subtype, _ = InvestmentSubType.objects.get_or_create(
                investment_type=renda_fixa_type,
                code='FUNDO_RF_PRE',
                defaults={'name': 'Fundos de Renda Fixa Prefixado', 'display_order': 5, 'is_active': True}
            )
            fund_ipca_subtype, _ = InvestmentSubType.objects.get_or_create(
                investment_type=renda_fixa_type,
                code='FUNDO_RF_IPCA',
                defaults={'name': 'Fundos de Renda Fixa IPCA+', 'display_order': 6, 'is_active': True}
            )
            
            # Map fund_type to InvestmentType/InvestmentSubType
            fund_type_map = {
                'RF_POS': (renda_fixa_type, fund_posfixado_subtype),
                'RF_PRE': (renda_fixa_type, fund_prefixado_subtype),
                'RF_IPCA': (renda_fixa_type, fund_ipca_subtype),
                # For other types, use Renda Fixa as default
                'MULTI': (renda_fixa_type, fund_posfixado_subtype),
                'ACOES': (renda_fixa_type, fund_posfixado_subtype),
                'CAMBIO': (renda_fixa_type, fund_posfixado_subtype),
                'OTHER': (renda_fixa_type, fund_posfixado_subtype),
            }
            
            # Extract and create CAIXA position from Saldo Disponível
            cash_balance = PortfolioExcelImportService.extract_cash_balance(worksheet)
            if cash_balance and cash_balance > 0:
                try:
                    # Create CAIXA position
                    caixa_asset_code = f"CAIXA_{user_id}"
                    caixa_asset_name = "XP Investimentos - Conta Investimento"
                    
                    # Use a fixed date (first day of current year) for CAIXA so it always updates the same position
                    today = timezone.now().date()
                    fixed_date = datetime(today.year, 1, 1).date()
                    # Set maturity far in the future for cash (or use a default)
                    future_date = datetime(today.year + 10, 12, 31).date()
                    
                    # Get or create Caixa subtype
                    caixa_subtype, _ = InvestmentSubType.objects.get_or_create(
                        investment_type=renda_fixa_type,
                        code='CAIXA',
                        defaults={
                            'name': 'Caixa',
                            'display_order': 1,
                            'is_predefined': False,
                            'is_active': True
                        }
                    )
                    
                    caixa_position, created = FixedIncomePosition.objects.update_or_create(
                        user_id=user_id,
                        asset_code=caixa_asset_code,
                        application_date=fixed_date,
                        defaults={
                            'asset_name': caixa_asset_name,
                            'maturity_date': future_date,
                            'quantity': Decimal('1.00'),
                            'available_quantity': Decimal('1.00'),
                            'guarantee_quantity': Decimal('0.00'),
                            'applied_value': cash_balance,
                            'position_value': cash_balance,
                            'net_value': cash_balance,
                            'gross_yield': Decimal('0.00'),
                            'net_yield': Decimal('0.00'),
                            'income_tax': Decimal('0.00'),
                            'iof': Decimal('0.00'),
                            'liquidity': 'Imediata',
                            'investment_type': renda_fixa_type,
                            'investment_sub_type': caixa_subtype,
                            'source': 'Excel Import',
                            'import_date': timezone.now(),
                        }
                    )
                    
                    if created:
                        results['created'] += 1
                    else:
                        results['updated'] += 1
                    results['caixa_count'] += 1
                    results['debug_info'].append(f"CAIXA position {'created' if created else 'updated'}: R$ {cash_balance}")
                except Exception as e:
                    results['errors'].append(f"Error creating CAIXA position: {str(e)}")
                    results['debug_info'].append(f"Failed to create CAIXA: {str(e)}")
            
            rows_processed = 0
            sections_found = []
            current_tesouro_subsection = None
            current_tesouro_subtype = None
            
            # Column mapping for Tesouro Direto (will be detected from header)
            tesouro_column_map = {}
            
            # Create or get Tesouro Direto sub-types under RENDA_FIXA
            # These are sub-types of TESOURO_DIRETO, but stored as separate sub-types under RENDA_FIXA
            # For now, we'll use the main TESOURO_DIRETO sub-type for all Tesouro positions
            # The specific sub-types (Pós-Fixado, Inflação, Prefixado) can be added later if needed
            
            # Map sub-type names to InvestmentSubType objects
            # Default to TESOURO_DIRETO sub-type for all Tesouro positions
            subtype_map = {
                'Pós-Fixado': tesouro_subtype,
                'Inflação': tesouro_subtype,
                'Prefixado': tesouro_subtype,
                'Tesouro Direto': tesouro_subtype,
                'TESOURO_DIRETO': tesouro_subtype,
            }
            
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                # Skip empty rows
                if not row or all(cell is None for cell in row):
                    continue
                
                rows_processed += 1
                
                # Detect section
                section = PortfolioExcelImportService.detect_section(row)
                if section:
                    current_section = section
                    sections_found.append(section)
                    current_tesouro_subsection = None
                    current_tesouro_subtype = None
                    current_fund_type = None
                    results['debug_info'].append(f"Row {row_idx}: Found section '{section}' - {str(row[0])[:100]}")
                    continue
                
                # Detect Investment Fund sub-section (e.g., "0,54%|Fundos de Renda Fixa Pós-Fixado")
                if current_section == 'fundos_investimento':
                    fund_type = PortfolioExcelImportService.detect_fund_type_subsection(row)
                    if fund_type:
                        current_fund_type = fund_type
                        results['debug_info'].append(f"Row {row_idx}: Found Fund sub-section '{fund_type}' - skipping header row")
                        continue  # Skip this row as it's the sub-section header
                
                # Detect Tesouro Direto sub-section (e.g., "15,2% | Pós-Fixado" or "2,2% | Inflação")
                # Note: In Excel, the sub-section header row also contains column headers
                # Also check in 'renda_fixa' section as Tesouro bonds can appear there
                if current_section in ['tesouro_direto', 'renda_fixa']:
                    subsection_info = PortfolioExcelImportService.detect_tesouro_subsection(row)
                    if subsection_info:
                        sub_type_name, allocation_pct = subsection_info
                        current_tesouro_subsection = sub_type_name
                        # Map to InvestmentSubType
                        current_tesouro_subtype = subtype_map.get(sub_type_name)
                        if not current_tesouro_subtype:
                            # Try to find by name match
                            for key, subtype_obj in subtype_map.items():
                                if key.lower() in sub_type_name.lower() or sub_type_name.lower() in key.lower():
                                    current_tesouro_subtype = subtype_obj
                                    break
                        results['debug_info'].append(f"Row {row_idx}: Found Tesouro sub-section '{sub_type_name}' ({allocation_pct}%) in section '{current_section}' - skipping header row")
                        continue  # Skip this row as it's both sub-section header and column headers
                
                # Skip header rows
                if PortfolioExcelImportService.is_header_row(row):
                    results['debug_info'].append(f"Row {row_idx}: Skipped header row")
                    continue
                
                # Process Renda Fixa entries (can be CDB or Tesouro Direto)
                if current_section == 'renda_fixa':
                    # First, try to extract Tesouro Direto bonds (LTN, NTN-B, LFT, etc.)
                    tesouro_data = PortfolioExcelImportService.extract_tesouro_from_row(
                        row, user_id, current_tesouro_subsection
                    )
                    if tesouro_data:
                        try:
                            titulo_name = tesouro_data.pop('titulo_name')
                            vencimento = tesouro_data.pop('vencimento')
                            sub_type_name = tesouro_data.pop('sub_type_name', None)
                            
                            # Use current sub-type if available, otherwise determine from bond
                            investment_sub_type = current_tesouro_subtype
                            if not investment_sub_type and sub_type_name:
                                investment_sub_type = subtype_map.get(sub_type_name)
                            
                            # For Tesouro positions, use only user_id and asset_code to prevent duplicates
                            existing_positions = FixedIncomePosition.objects.filter(
                                user_id=tesouro_data['user_id'],
                                asset_code=tesouro_data['asset_code']
                            ).order_by('-created_at')
                            
                            if existing_positions.exists():
                                position_list = list(existing_positions)
                                existing_position = position_list[0]
                                
                                duplicates_count = len(position_list) - 1
                                if duplicates_count > 0:
                                    for dup_position in position_list[1:]:
                                        try:
                                            dup_position.tesourodiretoposition.delete()
                                        except TesouroDiretoPosition.DoesNotExist:
                                            pass
                                        dup_position.delete()
                                    results['debug_info'].append(f"Row {row_idx}: Removed {duplicates_count} duplicate(s) for {tesouro_data['asset_code']}")
                                
                                for key, value in tesouro_data.items():
                                    setattr(existing_position, key, value)
                                existing_position.investment_type = renda_fixa_type
                                existing_position.investment_sub_type = investment_sub_type or tesouro_subtype
                                existing_position.source = 'Excel Import'
                                existing_position.import_date = timezone.now()
                                existing_position.save()
                                position = existing_position
                                created = False
                            else:
                                position, created = FixedIncomePosition.objects.update_or_create(
                                    user_id=tesouro_data['user_id'],
                                    asset_code=tesouro_data['asset_code'],
                                    defaults={
                                        **tesouro_data,
                                        'investment_type': renda_fixa_type,
                                        'investment_sub_type': investment_sub_type or tesouro_subtype,
                                    }
                                )
                            
                            # Create or update TesouroDiretoPosition
                            tesouro_pos, _ = TesouroDiretoPosition.objects.update_or_create(
                                fixed_income_position=position,
                                defaults={
                                    'titulo_name': titulo_name,
                                    'vencimento': vencimento,
                                }
                            )
                            
                            if created:
                                results['created'] += 1
                            else:
                                results['updated'] += 1
                            results['tesouro_count'] += 1
                            results['debug_info'].append(f"Row {row_idx}: {'Created' if created else 'Updated'} Tesouro '{titulo_name[:50]}' in Renda Fixa section")
                        except Exception as e:
                            results['errors'].append(f"Row {row_idx}: {str(e)}")
                            import traceback
                            results['debug_info'].append(f"Row {row_idx}: Error details: {traceback.format_exc()}")
                    else:
                        # Skip rows that are proventos, dividendos, distribuições, etc.
                        first_cell = str(row[0]).strip() if row[0] else ''
                        first_cell_upper = first_cell.upper()
                        excluded_keywords = [
                            'PROVENTOS', 'DIVIDENDOS', 'DISTRIBUIÇÕES', 'DISTRIBUICOES',
                            'RENDIMENTOS', 'JUROS', 'AMORTIZAÇÃO', 'AMORTIZACAO',
                            'RESGATE', 'LIQUIDAÇÃO', 'LIQUIDACAO',
                            'DIVIDENDOS, PROVENTOS', 'PROVENTOS E OUTRAS DISTRIBUIÇÕES'
                        ]
                        
                        if any(keyword in first_cell_upper for keyword in excluded_keywords):
                            results['debug_info'].append(f"Row {row_idx}: Skipped non-investment entry in Renda Fixa: '{first_cell[:50]}'")
                            continue
                        
                        # If not Tesouro, try CDB
                        cdb_data = PortfolioExcelImportService.extract_cdb_from_row(row, user_id, current_section)
                        if cdb_data:
                            try:
                                position, created = FixedIncomePosition.objects.update_or_create(
                                    user_id=cdb_data['user_id'],
                                    asset_code=cdb_data['asset_code'],
                                    application_date=cdb_data['application_date'],
                                    defaults={
                                        **cdb_data,
                                        'investment_type': renda_fixa_type,
                                        'investment_sub_type': cdb_prefixado_subtype,  # Assign CDB_PREFIXADO sub-type
                                    }
                                )
                                
                                if created:
                                    results['created'] += 1
                                else:
                                    results['updated'] += 1
                                results['cdb_count'] += 1
                            except Exception as e:
                                results['errors'].append(f"Row {row_idx}: {str(e)}")
                        else:
                            # Log why extraction failed
                            if row[0]:
                                asset_name = str(row[0]).strip()
                                if asset_name:
                                    results['debug_info'].append(f"Row {row_idx}: Skipped (not CDB or Tesouro): {asset_name[:50]}")
                
                # Process Tesouro Direto entries
                elif current_section == 'tesouro_direto':
                    tesouro_data = PortfolioExcelImportService.extract_tesouro_from_row(
                        row, user_id, current_tesouro_subsection
                    )
                    if tesouro_data:
                        try:
                            titulo_name = tesouro_data.pop('titulo_name')
                            vencimento = tesouro_data.pop('vencimento')
                            sub_type_name = tesouro_data.pop('sub_type_name', None)
                            
                            # Use current sub-type if available, otherwise determine from bond
                            investment_sub_type = current_tesouro_subtype
                            if not investment_sub_type and sub_type_name:
                                investment_sub_type = subtype_map.get(sub_type_name)
                            
                            # For Tesouro positions, use only user_id and asset_code to prevent duplicates
                            # The same bond (e.g., "LFT mar/2029") should always update the same position
                            # regardless of application_date or which subsection it appears in
                            existing_positions = FixedIncomePosition.objects.filter(
                                user_id=tesouro_data['user_id'],
                                asset_code=tesouro_data['asset_code']
                            ).order_by('-created_at')
                            
                            if existing_positions.exists():
                                # If duplicates exist, keep the most recent one and delete others
                                position_list = list(existing_positions)
                                existing_position = position_list[0]
                                
                                # Delete any duplicates (keep only the first/most recent)
                                duplicates_count = len(position_list) - 1
                                if duplicates_count > 0:
                                    # Delete all except the first one
                                    for dup_position in position_list[1:]:
                                        # Also delete associated TesouroDiretoPosition if exists
                                        try:
                                            dup_position.tesourodiretoposition.delete()
                                        except TesouroDiretoPosition.DoesNotExist:
                                            pass
                                        dup_position.delete()
                                    results['debug_info'].append(f"Row {row_idx}: Removed {duplicates_count} duplicate(s) for {tesouro_data['asset_code']}")
                                
                                # Update existing position with new data
                                for key, value in tesouro_data.items():
                                    setattr(existing_position, key, value)
                                existing_position.investment_type = renda_fixa_type
                                existing_position.investment_sub_type = investment_sub_type or tesouro_subtype
                                existing_position.source = 'Excel Import'
                                existing_position.import_date = timezone.now()
                                existing_position.save()
                                position = existing_position
                                created = False
                            else:
                                # Create new position using only user_id and asset_code as unique identifiers
                                # This ensures the same bond always updates the same position
                                position, created = FixedIncomePosition.objects.update_or_create(
                                    user_id=tesouro_data['user_id'],
                                    asset_code=tesouro_data['asset_code'],
                                    defaults={
                                        **tesouro_data,
                                        'investment_type': renda_fixa_type,  # RENDA_FIXA, not Tesouro Direto
                                        'investment_sub_type': investment_sub_type or tesouro_subtype,  # TESOURO_DIRETO sub-type
                                    }
                                )
                            
                            # Create or update TesouroDiretoPosition
                            tesouro_pos, _ = TesouroDiretoPosition.objects.update_or_create(
                                fixed_income_position=position,
                                defaults={
                                    'titulo_name': titulo_name,
                                    'vencimento': vencimento,
                                }
                            )
                            
                            if created:
                                results['created'] += 1
                            else:
                                results['updated'] += 1
                            results['tesouro_count'] += 1
                        except Exception as e:
                            results['errors'].append(f"Row {row_idx}: {str(e)}")
                            import traceback
                            results['debug_info'].append(f"Row {row_idx}: Error details: {traceback.format_exc()}")
                    else:
                        # Log why Tesouro extraction failed
                        if row[0]:
                            titulo_name = str(row[0]).strip()
                            if titulo_name and not any(bond_type in titulo_name.upper() for bond_type in ['LFT', 'NTNB', 'LTN', 'NTN-B', 'NTN B', 'TESOURO']):
                                results['debug_info'].append(f"Row {row_idx}: Skipped (not Tesouro): {titulo_name[:50]}")
                # Process Investment Funds entries
                elif current_section == 'fundos_investimento':
                    fund_type_code = current_fund_type or 'RF_POS'
                    fund_data = PortfolioExcelImportService.extract_investment_fund_from_row(
                        row, user_id, fund_type_code
                    )
                    if fund_data:
                        # Additional validation: ensure it's a real fund
                        fund_name = fund_data.get('fund_name', '')
                        if not fund_name or len(fund_name.strip()) < 3:
                            results['debug_info'].append(f"Row {row_idx}: Skipped fund (name too short or empty): '{fund_name}'")
                            continue
                        try:
                            # Map fund_type to InvestmentType and InvestmentSubType
                            investment_type, investment_sub_type = fund_type_map.get(
                                fund_type_code, 
                                (renda_fixa_type, fund_posfixado_subtype)
                            )
                            fund_data['investment_type'] = investment_type
                            fund_data['investment_sub_type'] = investment_sub_type
                            
                            # Use only user_id and fund_name as unique key (not quota_date)
                            # If fund exists, UPDATE values (replace, not aggregate) - same fund with different quota_date
                            # When same fund appears with different quota_date, update: quota_value, position_value, net_value
                            try:
                                existing_fund = InvestmentFund.objects.get(
                                    user_id=fund_data['user_id'],
                                    fund_name=fund_data['fund_name']
                                )
                                
                                # Always update quota_date and quota_value to the latest (newer date)
                                if fund_data.get('quota_date'):
                                    # Normalize dates to date objects for comparison
                                    new_quota_date = fund_data['quota_date']
                                    if isinstance(new_quota_date, datetime):
                                        new_quota_date = new_quota_date.date()
                                    elif not isinstance(new_quota_date, date):
                                        # If it's not a date or datetime, skip
                                        new_quota_date = None
                                    
                                    if new_quota_date:
                                        existing_quota_date = existing_fund.quota_date
                                        if isinstance(existing_quota_date, datetime):
                                            existing_quota_date = existing_quota_date.date()
                                        
                                        # Update if no existing date or new date is >= existing date
                                        if not existing_quota_date or new_quota_date >= existing_quota_date:
                                            existing_fund.quota_date = new_quota_date
                                            if fund_data.get('quota_value') is not None:
                                                existing_fund.quota_value = fund_data.get('quota_value')
                                
                                # REPLACE these values with new ones (not sum) - they represent current state at the quota_date
                                if fund_data.get('quota_quantity') is not None:
                                    existing_fund.quota_quantity = fund_data.get('quota_quantity')
                                if fund_data.get('in_quotation') is not None:
                                    existing_fund.in_quotation = fund_data.get('in_quotation')
                                if fund_data.get('position_value') is not None:
                                    existing_fund.position_value = fund_data.get('position_value')
                                if fund_data.get('net_value') is not None:
                                    existing_fund.net_value = fund_data.get('net_value')
                                if fund_data.get('applied_value') is not None:
                                    existing_fund.applied_value = fund_data.get('applied_value')
                                
                                # Recalculate returns based on updated values
                                if existing_fund.applied_value and existing_fund.applied_value > 0 and existing_fund.net_value and existing_fund.net_value > 0:
                                    existing_fund.net_return_percent = ((existing_fund.net_value - existing_fund.applied_value) / existing_fund.applied_value) * 100
                                
                                # Update other fields from fund_data
                                for key, value in fund_data.items():
                                    if key not in ['user_id', 'fund_name', 'quota_quantity', 'in_quotation', 'position_value', 'net_value', 'applied_value', 'quota_date', 'quota_value', 'net_return_percent', 'gross_return_percent']:
                                        if value is not None:
                                            setattr(existing_fund, key, value)
                                
                                existing_fund.save()
                                results['updated'] += 1
                                results['debug_info'].append(f"Row {row_idx}: Updated fund '{fund_data['fund_name'][:50]}' (quota_date: {fund_data.get('quota_date')})")
                                
                            except InvestmentFund.DoesNotExist:
                                # Create new fund
                                fund = InvestmentFund.objects.create(**fund_data)
                                results['created'] += 1
                                results['debug_info'].append(f"Row {row_idx}: Created fund '{fund_data['fund_name'][:50]}'")
                            
                            # Track fund count
                            if 'fund_count' not in results:
                                results['fund_count'] = 0
                            results['fund_count'] += 1
                        except Exception as e:
                            results['errors'].append(f"Row {row_idx}: {str(e)}")
                    else:
                        # Log why fund extraction failed
                        if row[0]:
                            fund_name = str(row[0]).strip()
                            if fund_name and len(fund_name) > 3:
                                # Check if it was skipped due to excluded keywords
                                fund_name_upper = fund_name.upper()
                                excluded_keywords = ['PROVENTOS', 'DIVIDENDOS', 'DISTRIBUIÇÕES', 'DISTRIBUICOES', 'RENDIMENTOS']
                                if any(keyword in fund_name_upper for keyword in excluded_keywords):
                                    results['debug_info'].append(f"Row {row_idx}: Skipped non-fund entry (proventos/dividendos): {fund_name[:50]}")
                                else:
                                    results['debug_info'].append(f"Row {row_idx}: Skipped fund row (extraction failed): {fund_name[:50]}")
                
                else:
                    # Log rows that don't match any section
                    if row[0] and current_section is None:
                        first_cell = str(row[0]).strip()[:50]
                        if first_cell:
                            results['debug_info'].append(f"Row {row_idx}: No section detected, first cell: {first_cell}")
            
            results['debug_info'].append(f"Total rows processed: {rows_processed}")
            results['debug_info'].append(f"Sections found: {', '.join(sections_found) if sections_found else 'None'}")
            
            if not sections_found:
                results['errors'].append("Nenhuma seção 'RENDA FIXA', 'TESOURO DIRETO' ou 'FUNDOS DE INVESTIMENTO' foi encontrada no arquivo. Verifique o formato do arquivo Excel.")
        
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            results['errors'].append(f"Import error: {str(e)}")
            results['error_details'] = error_details
        
        return results


