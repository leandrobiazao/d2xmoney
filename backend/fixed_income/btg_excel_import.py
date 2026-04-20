"""
BTG Pactual Posição consolidada Excel import (abas Renda Fixa + Conta Corrente).

Secção «Posições Detalhadas» → blocos «Detalhamento > …» (Tesouro/títulos públicos).
CDB/LCI noutras linhas do mesmo export: não tratado neste parser; pode ser estendido
com deteção por cabeçalho/célula «CDB» no futuro.
"""
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from django.utils import timezone

from configuration.models import InvestmentType, InvestmentSubType
from .models import FixedIncomePosition, TesouroDiretoPosition


def _portfolio_service():
    """Import after fixed_income.services is fully loaded (avoids circular import)."""
    from .services import PortfolioExcelImportService
    return PortfolioExcelImportService


def _norm(s: str) -> str:
    if not s:
        return ''
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()


def find_sheet_by_title(workbook: openpyxl.Workbook, *candidates: str):
    for sheet in workbook.worksheets:
        for c in candidates:
            if _norm(sheet.title) == _norm(c):
                return sheet
    return None


def _cell_str(val: Any) -> str:
    if val is None:
        return ''
    if isinstance(val, float):
        return str(val)
    return str(val).strip()


def _parse_any_date(val: Any) -> Optional[date]:
    if val is None or val == '' or val == '-':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    return _portfolio_service().parse_date(s)


def _parse_any_currency(val: Any) -> Decimal:
    if val is None or val == '' or val == '-':
        return Decimal('0.00')
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    return _portfolio_service().parse_currency(str(val))


def _find_posicoes_detalhadas_row(ws) -> Optional[int]:
    """1-based row index of marker cell, or None."""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        for cell in row:
            t = _cell_str(cell)
            if 'posicoes detalhadas' in _norm(t) or 'posições detalhadas' in t.lower():
                return row_idx
    return None


def _is_detalhamento_title(row: Tuple) -> bool:
    b = _cell_str(row[1]) if len(row) > 1 else ''
    return 'detalhamento' in _norm(b) and '>' in b


def _row_looks_like_header(row: Tuple) -> bool:
    joined = ' '.join(_cell_str(c) for c in row[:20])
    n = _norm(joined)
    return 'ativo' in n and 'vencimento' in n


def _map_header(row: Tuple) -> Dict[str, int]:
    """Map normalized header label -> column index (0-based)."""
    idx_map: Dict[str, int] = {}
    for i, cell in enumerate(row):
        key = _norm(_cell_str(cell))
        if not key or key == 'none':
            continue
        key = key.replace('r$', '').strip()
        if 'saldo liquido' in key or 'saldo líquido' in _cell_str(cell).lower():
            idx_map['saldo_liquido'] = i
        elif 'saldo bruto' in key:
            idx_map['saldo_bruto'] = i
        elif key == 'ativo' or key.startswith('ativo'):
            idx_map['ativo'] = i
        elif 'emissao' in key or 'emissão' in _cell_str(cell).lower():
            idx_map['emissao'] = i
        elif 'vencimento' in key:
            idx_map['vencimento'] = i
        elif 'aquisicao' in key or 'aquisição' in _cell_str(cell).lower():
            idx_map['aquisicao'] = i
        elif 'liquidez' in key and 'dias' not in key and 'inicial' not in key:
            idx_map['liquidez'] = i
        elif 'taxa' in key and 'compra' in key:
            idx_map['taxa_compra'] = i
        elif key == 'quantidade' or key.startswith('quantidade'):
            idx_map['quantidade'] = i
        elif 'preco compra' in key or 'preço compra' in _cell_str(cell).lower():
            idx_map['preco_compra'] = i
        elif 'valor compra' in key:
            idx_map['valor_compra'] = i
        elif ('preco' in key or 'preço' in _cell_str(cell).lower()) and 'compra' not in key and 'saldo' not in key:
            if 'preco' == key or _norm(_cell_str(cell)) in ('preco r$', 'preço r$'):
                idx_map['preco'] = i
        elif key.startswith('ir') and 'r$' in _norm(_cell_str(cell)):
            idx_map['ir'] = i
        elif 'iof' in key:
            idx_map['iof'] = i
    return idx_map


def _btg_asset_code(tipo: str, vencimento: date) -> str:
    t = re.sub(r'[^A-Z0-9]', '_', tipo.strip().upper())
    return f'BTG_{t}_{vencimento.strftime("%Y%m%d")}'


def extract_btg_conta_corrente_balance(ws) -> Optional[Decimal]:
    """Valor financeiro R$ block: header row then data row."""
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            cs = _cell_str(cell)
            if 'valor financeiro' in _norm(cs):
                # value typically same row+1, col j+1 or j
                if i + 1 < len(rows):
                    nxt = rows[i + 1]
                    if j < len(nxt) and nxt[j] is not None:
                        v = _parse_any_currency(nxt[j])
                        if v and v > 0:
                            return v
                    if j + 1 < len(nxt) and nxt[j + 1] is not None:
                        v = _parse_any_currency(nxt[j + 1])
                        if v and v > 0:
                            return v
    return None


def import_btg_excel(file_path: str, user_id: str) -> Dict:
    results = {
        'created': 0,
        'updated': 0,
        'errors': [],
        'cdb_count': 0,
        'tesouro_count': 0,
        'caixa_count': 0,
        'debug_info': [],
    }

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws_rf = find_sheet_by_title(wb, 'Renda Fixa', 'RENDA FIXA')
    ws_cc = find_sheet_by_title(wb, 'Conta Corrente', 'CONTA CORRENTE')
    if not ws_rf:
        results['errors'].append("Aba 'Renda Fixa' não encontrada no Excel BTG.")
        return results
    results['debug_info'].append(f"BTG: aba Renda Fixa = {ws_rf.title}")

    renda_fixa_type, _ = InvestmentType.objects.get_or_create(
        code='RENDA_FIXA',
        defaults={'name': 'Renda Fixa', 'display_order': 2}
    )
    tesouro_subtype, _ = InvestmentSubType.objects.get_or_create(
        investment_type=renda_fixa_type,
        code='TESOURO_DIRETO',
        defaults={'name': 'Tesouro Direto', 'display_order': 2, 'is_active': True}
    )
    caixa_subtype, _ = InvestmentSubType.objects.get_or_create(
        investment_type=renda_fixa_type,
        code='CAIXA',
        defaults={
            'name': 'Caixa',
            'display_order': 1,
            'is_predefined': False,
            'is_active': True
        }
    )

    # Conta Corrente -> CAIXA
    if ws_cc:
        cash = extract_btg_conta_corrente_balance(ws_cc)
        results['debug_info'].append(f"BTG: aba Conta Corrente = {ws_cc.title}, saldo={cash}")
        if cash is not None and cash >= 0:
            try:
                caixa_asset_code = f'CAIXA_{user_id}'
                today = timezone.now().date()
                fixed_date = datetime(today.year, 1, 1).date()
                future_date = datetime(today.year + 10, 12, 31).date()
                FixedIncomePosition.objects.filter(
                    user_id=user_id,
                    asset_code__startswith='CAIXA_'
                ).delete()
                caixa_position, created = FixedIncomePosition.objects.update_or_create(
                    user_id=user_id,
                    asset_code=caixa_asset_code,
                    application_date=fixed_date,
                    defaults={
                        'asset_name': 'BTG Pactual - Conta Corrente (saldo disponível)',
                        'maturity_date': future_date,
                        'quantity': Decimal('1.00'),
                        'available_quantity': Decimal('1.00'),
                        'guarantee_quantity': Decimal('0.00'),
                        'applied_value': cash,
                        'position_value': cash,
                        'net_value': cash,
                        'gross_yield': Decimal('0.00'),
                        'net_yield': Decimal('0.00'),
                        'income_tax': Decimal('0.00'),
                        'iof': Decimal('0.00'),
                        'liquidity': 'Imediata',
                        'investment_type': renda_fixa_type,
                        'investment_sub_type': caixa_subtype,
                        'source': 'Excel Import',
                        'import_date': timezone.now(),
                    }
                )
                results['created' if created else 'updated'] += 1
                results['caixa_count'] += 1
            except Exception as e:
                results['errors'].append(f"CAIXA BTG: {e}")
    else:
        results['debug_info'].append("BTG: aba Conta Corrente não encontrada — CAIXA não atualizada")

    marker_row = _find_posicoes_detalhadas_row(ws_rf)
    if not marker_row:
        results['errors'].append(
            "Marcador «Posições Detalhadas» não encontrado na aba Renda Fixa."
        )
        return results
    results['debug_info'].append(f"BTG: «Posições Detalhadas» na linha {marker_row}")

    rows = list(ws_rf.iter_rows(values_only=True))
    # Conteúdo importável começa na linha seguinte ao título «Posições Detalhadas»
    start_idx = marker_row + 1
    header_map: Optional[Dict[str, int]] = None
    in_detalhamento = False

    for global_idx in range(start_idx, len(rows) + 1):
        row = rows[global_idx - 1]
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        if _is_detalhamento_title(row):
            in_detalhamento = True
            header_map = None
            continue

        if not in_detalhamento:
            continue

        if _row_looks_like_header(row):
            header_map = _map_header(row)
            results['debug_info'].append(f"BTG: cabeçalho detectado linha {global_idx}: {header_map}")
            continue

        if not header_map:
            continue

        ativo_i = header_map.get('ativo')
        venc_i = header_map.get('vencimento')
        if ativo_i is None or venc_i is None:
            continue

        tipo = _cell_str(row[ativo_i]) if ativo_i < len(row) else ''
        if not tipo or _norm(tipo) == 'total':
            continue

        v_raw = row[venc_i] if venc_i < len(row) else None
        vencimento = _parse_any_date(v_raw)
        if not vencimento:
            continue

        asset_code = _btg_asset_code(tipo, vencimento)
        titulo_name = f"{tipo.strip()} {vencimento.strftime('%d/%m/%Y')}"

        applied = _parse_any_currency(row[header_map['valor_compra']] if 'valor_compra' in header_map and header_map['valor_compra'] < len(row) else None)
        saldo_bruto = _parse_any_currency(row[header_map['saldo_bruto']] if 'saldo_bruto' in header_map and header_map['saldo_bruto'] < len(row) else None)
        ir_ = _parse_any_currency(row[header_map['ir']] if 'ir' in header_map and header_map['ir'] < len(row) else None)
        iof_ = _parse_any_currency(row[header_map['iof']] if 'iof' in header_map and header_map['iof'] < len(row) else None)

        if 'saldo_liquido' in header_map and header_map['saldo_liquido'] < len(row):
            saldo_liq = _parse_any_currency(row[header_map['saldo_liquido']])
        else:
            saldo_liq = Decimal('0.00')
        if saldo_liq and saldo_liq > 0:
            net_val = saldo_liq
        else:
            net_val = (saldo_bruto - ir_ - iof_) if saldo_bruto else Decimal('0.00')

        qty = _portfolio_service().parse_quantity(
            str(row[header_map['quantidade']]) if 'quantidade' in header_map and header_map['quantidade'] < len(row) else '0'
        )
        preco = _parse_any_currency(row[header_map['preco']] if 'preco' in header_map and header_map['preco'] < len(row) else None)

        rate = None
        if 'taxa_compra' in header_map and header_map['taxa_compra'] < len(row):
            r = row[header_map['taxa_compra']]
            rate = _cell_str(r) if r is not None else None
            if rate and len(rate) > 20:
                rate = rate[:20]

        liq = None
        if 'liquidez' in header_map and header_map['liquidez'] < len(row):
            liq = _cell_str(row[header_map['liquidez']])

        app_date = _parse_any_date(row[header_map['aquisicao']] if 'aquisicao' in header_map and header_map['aquisicao'] < len(row) else None)
        if not app_date:
            app_date = date(vencimento.year, 1, 1)

        gross_yield = saldo_bruto - applied if saldo_bruto and applied else Decimal('0.00')
        net_yield = net_val - applied if net_val and applied else Decimal('0.00')

        data = {
            'user_id': user_id,
            'asset_name': titulo_name,
            'asset_code': asset_code,
            'application_date': app_date,
            'grace_period_end': None,
            'maturity_date': vencimento,
            'price_date': None,
            'rate': rate,
            'price': preco,
            'quantity': qty,
            'available_quantity': qty,
            'guarantee_quantity': Decimal('0.00'),
            'applied_value': applied,
            'position_value': saldo_bruto,
            'net_value': net_val,
            'gross_yield': gross_yield,
            'net_yield': net_yield,
            'income_tax': ir_,
            'iof': iof_,
            'liquidity': liq,
            'source': 'Excel Import',
            'import_date': timezone.now(),
        }

        try:
            qs = FixedIncomePosition.objects.filter(
                user_id=user_id,
                asset_code=asset_code,
            ).order_by('-updated_at')
            lst = list(qs)
            if len(lst) > 1:
                for dup in lst[1:]:
                    try:
                        dup.tesouro_direto.delete()
                    except TesouroDiretoPosition.DoesNotExist:
                        pass
                    dup.delete()
                results['debug_info'].append(f"BTG: removidos duplicados {asset_code}")
            lst = list(
                FixedIncomePosition.objects.filter(
                    user_id=user_id,
                    asset_code=asset_code,
                ).order_by('-updated_at')
            )
            if lst:
                position = lst[0]
                for k, v in data.items():
                    setattr(position, k, v)
                position.investment_type = renda_fixa_type
                position.investment_sub_type = tesouro_subtype
                position.save()
                created = False
            else:
                position = FixedIncomePosition.objects.create(
                    investment_type=renda_fixa_type,
                    investment_sub_type=tesouro_subtype,
                    **data,
                )
                created = True

            TesouroDiretoPosition.objects.update_or_create(
                fixed_income_position=position,
                defaults={'titulo_name': titulo_name, 'vencimento': vencimento},
            )
            if created:
                results['created'] += 1
            else:
                results['updated'] += 1
            results['tesouro_count'] += 1
        except Exception as e:
            results['errors'].append(f"Linha {global_idx} ({asset_code}): {e}")

    if results['tesouro_count'] == 0 and not results['errors']:
        results['errors'].append(
            'Nenhuma posição importada após «Posições Detalhadas». Verifique cabeçalhos (Ativo, Vencimento).'
        )

    return results


def user_is_btg(account_provider: str) -> bool:
    if not account_provider:
        return False
    return 'btg' in account_provider.lower()
