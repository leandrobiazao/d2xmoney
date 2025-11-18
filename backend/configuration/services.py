"""
Service for managing investment types and sub-types configuration.
"""
import openpyxl
from typing import List, Dict, Optional
from django.db import transaction
from .models import InvestmentType, InvestmentSubType


class ConfigurationService:
    """Service for managing investment types and sub-types."""
    
    @staticmethod
    def get_all_investment_types(active_only: bool = True) -> List[InvestmentType]:
        """Get all investment types."""
        queryset = InvestmentType.objects.all()
        if active_only:
            queryset = queryset.filter(is_active=True)
        return list(queryset.order_by('display_order', 'name'))
    
    @staticmethod
    def get_investment_type_by_code(code: str) -> Optional[InvestmentType]:
        """Get investment type by code."""
        try:
            return InvestmentType.objects.get(code=code, is_active=True)
        except InvestmentType.DoesNotExist:
            return None
    
    @staticmethod
    def get_sub_types_by_investment_type(
        investment_type: InvestmentType,
        active_only: bool = True
    ) -> List[InvestmentSubType]:
        """Get all sub-types for an investment type."""
        queryset = InvestmentSubType.objects.filter(investment_type=investment_type)
        if active_only:
            queryset = queryset.filter(is_active=True)
        return list(queryset.order_by('display_order', 'name'))
    
    @staticmethod
    @transaction.atomic
    def import_sub_types_from_excel(
        file_path: str,
        investment_type_code: str,
        sheet_name: Optional[str] = None
    ) -> Dict[str, int]:
        """
        Import sub-types from Excel file.
        
        Expected Excel format:
        - Column A: Sub-type name
        - Column B: Sub-type code
        - Column C: Display order (optional)
        
        Returns dict with 'created', 'updated', 'errors' counts.
        """
        investment_type = ConfigurationService.get_investment_type_by_code(investment_type_code)
        if not investment_type:
            raise ValueError(f"Investment type with code '{investment_type_code}' not found")
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        
        created = 0
        updated = 0
        errors = []
        
        # Skip header row if present
        start_row = 2 if sheet.cell(1, 1).value and 'name' in str(sheet.cell(1, 1).value).lower() else 1
        
        for row_idx in range(start_row, sheet.max_row + 1):
            try:
                name = sheet.cell(row_idx, 1).value
                code = sheet.cell(row_idx, 2).value
                display_order = sheet.cell(row_idx, 3).value if sheet.cell(row_idx, 3).value else row_idx - start_row + 1
                
                if not name or not code:
                    continue
                
                name = str(name).strip()
                code = str(code).strip()
                
                sub_type, created_flag = InvestmentSubType.objects.update_or_create(
                    investment_type=investment_type,
                    code=code,
                    defaults={
                        'name': name,
                        'display_order': int(display_order) if display_order else row_idx - start_row + 1,
                        'is_predefined': True,
                        'is_active': True
                    }
                )
                
                if created_flag:
                    created += 1
                else:
                    updated += 1
                    
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        return {
            'created': created,
            'updated': updated,
            'errors': errors,
            'total_processed': created + updated
        }
    
    @staticmethod
    def create_investment_type(
        name: str,
        code: str,
        display_order: int = 0
    ) -> InvestmentType:
        """Create a new investment type."""
        return InvestmentType.objects.create(
            name=name,
            code=code,
            display_order=display_order,
            is_active=True
        )
    
    @staticmethod
    def create_sub_type(
        investment_type: InvestmentType,
        name: str,
        code: str,
        display_order: int = 0,
        is_predefined: bool = False
    ) -> InvestmentSubType:
        """Create a new sub-type."""
        return InvestmentSubType.objects.create(
            investment_type=investment_type,
            name=name,
            code=code,
            display_order=display_order,
            is_predefined=is_predefined,
            is_active=True
        )


